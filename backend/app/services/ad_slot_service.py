import csv
import io
import json
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func, select
from sqlalchemy.orm import Session as OrmSession

from ..db.models import AdCampaign, AdClick, AdImpression, AdSlot, User
from .ad_slot_helpers import (
    DEFAULT_AD_SLOTS,
    _campaign_image_formula_url,
    _campaign_matches_filters,
    _campaign_runtime_status,
    _campaign_runtime_variant,
    _ensure_campaign_links_for_row,
    _normalize_slot_config,
    _parse_optional_datetime,
    _safe_xlsx_filename,
    _serialize_slot,
)
from .user_region_service import sync_user_region_context as sync_user_region_context_record


def ensure_default_ad_slots(db: OrmSession) -> None:
    existing = {row.slot_key: row for row in db.scalars(select(AdSlot)).all()}
    dirty = False
    for slot in DEFAULT_AD_SLOTS:
        if slot['slotKey'] not in existing:
            db.add(
                AdSlot(
                    slot_key=slot['slotKey'],
                    placement_type=slot['placementType'],
                    status=slot['status'],
                    config_json=json.dumps(slot['config'], ensure_ascii=False),
                )
            )
            dirty = True
    if dirty:
        db.commit()


def _defaults_by_key() -> Dict[str, Dict[str, Any]]:
    return {slot['slotKey']: slot for slot in DEFAULT_AD_SLOTS}


AUDIENCE_TYPES = {'all', 'member', 'guest'}
REGION_LEVELS = {'all', 'city', 'district', 'neighborhood'}
REGION_LEVEL_SCORES = {'all': 0, 'city': 1, 'district': 2, 'neighborhood': 3}


@dataclass
class AdRuntimeContext:
    audience_type: str = 'all'
    city: Optional[str] = None
    district: Optional[str] = None
    neighborhood: Optional[str] = None


def _clean_region_string(value: Any) -> Optional[str]:
    if value is None:
        return None
    cleaned = str(value).strip()
    if not cleaned:
        return None
    return ' '.join(cleaned.split())


def _normalize_audience_type(value: Any) -> str:
    cleaned = str(value or '').strip().lower()
    aliases = {
        '': 'all',
        'all': 'all',
        '전체': 'all',
        'member': 'member',
        '회원': 'member',
        'guest': 'guest',
        '게스트': 'guest',
    }
    return aliases.get(cleaned, 'all')


def _normalize_region_level(value: Any) -> str:
    cleaned = str(value or '').strip().lower()
    aliases = {
        '': 'all',
        'all': 'all',
        '전체': 'all',
        'city': 'city',
        '시': 'city',
        'district': 'district',
        '구': 'district',
        'neighborhood': 'neighborhood',
        '동': 'neighborhood',
    }
    return aliases.get(cleaned, 'all')


def _resolve_region_label(city: Optional[str], district: Optional[str], neighborhood: Optional[str]) -> Optional[str]:
    primary = neighborhood or district
    if primary and city:
        return f'{primary}, {city}'
    return primary or city


def _resolve_target_region_level(
    level: Any,
    city: Optional[str],
    district: Optional[str],
    neighborhood: Optional[str],
) -> str:
    normalized = _normalize_region_level(level)
    if normalized != 'all':
        return normalized
    if neighborhood:
        return 'neighborhood'
    if district:
        return 'district'
    if city:
        return 'city'
    return 'all'


def _region_key_from_values(
    level: str,
    city: Optional[str],
    district: Optional[str] = None,
    neighborhood: Optional[str] = None,
) -> Optional[str]:
    normalized_level = _normalize_region_level(level)
    normalized_city = _clean_region_string(city)
    normalized_district = _clean_region_string(district)
    normalized_neighborhood = _clean_region_string(neighborhood)
    if normalized_level == 'all':
        return None
    if normalized_level == 'city':
        return f'city:{normalized_city}' if normalized_city else None
    if normalized_level == 'district':
        return f'district:{normalized_city}/{normalized_district or "_"}' if normalized_city else None
    if normalized_level == 'neighborhood':
        return f'neighborhood:{normalized_city}/{normalized_district or "_"}/{normalized_neighborhood or "_"}' if normalized_city else None
    return None


def _parse_region_key(value: Any) -> Optional[tuple[str, Optional[str], Optional[str], Optional[str]]]:
    raw = str(value or '').strip()
    if not raw or raw == 'all' or ':' not in raw:
        return None
    level_raw, remainder = raw.split(':', 1)
    level = _normalize_region_level(level_raw)
    if level == 'city':
        city = _clean_region_string(remainder)
        return (level, city, None, None) if city else None
    if level == 'district':
        city_raw, _, district_raw = remainder.partition('/')
        city = _clean_region_string(city_raw)
        district = _clean_region_string(district_raw)
        if district == '_':
            district = None
        return (level, city, district, None) if city else None
    if level == 'neighborhood':
        parts = remainder.split('/')
        if len(parts) < 3:
            return None
        city = _clean_region_string(parts[0])
        district = _clean_region_string(parts[1])
        neighborhood = _clean_region_string('/'.join(parts[2:]))
        if district == '_':
            district = None
        if neighborhood == '_':
            neighborhood = None
        return (level, city, district, neighborhood) if city else None
    return None


def _normalize_region_keys(value: Any) -> List[str]:
    if value is None:
        raw_values: List[Any] = []
    elif isinstance(value, (list, tuple, set)):
        raw_values = list(value)
    else:
        raw_values = []
        for chunk in str(value).replace(';', ',').split(','):
            raw_values.extend(part for part in chunk.splitlines())
    normalized: List[str] = []
    seen: set[str] = set()
    for raw in raw_values:
        parsed = _parse_region_key(raw)
        if not parsed:
            continue
        key = _region_key_from_values(parsed[0], parsed[1], parsed[2], parsed[3])
        if not key or key in seen:
            continue
        seen.add(key)
        normalized.append(key)
    return sorted(normalized)


def _region_keys_from_legacy_fields(
    level: str,
    city: Optional[str],
    district: Optional[str],
    neighborhood: Optional[str],
) -> List[str]:
    key = _region_key_from_values(level, city, district, neighborhood)
    return [key] if key else []


def _campaign_audience_type(campaign: AdCampaign) -> str:
    return _normalize_audience_type(getattr(campaign, 'audience_type', None))


def _campaign_region_level(campaign: AdCampaign) -> str:
    return _resolve_target_region_level(
        getattr(campaign, 'target_region_level', None),
        _clean_region_string(getattr(campaign, 'target_city', None)),
        _clean_region_string(getattr(campaign, 'target_district', None)),
        _clean_region_string(getattr(campaign, 'target_neighborhood', None)),
    )


def _campaign_region_keys(campaign: AdCampaign) -> List[str]:
    stored = getattr(campaign, 'target_region_keys_json', None)
    if stored and isinstance(stored, str):
        try:
            parsed = json.loads(stored)
        except Exception:
            parsed = None
        keys = _normalize_region_keys(parsed)
        if keys:
            return keys
    return _region_keys_from_legacy_fields(
        _campaign_region_level(campaign),
        getattr(campaign, 'target_city', None),
        getattr(campaign, 'target_district', None),
        getattr(campaign, 'target_neighborhood', None),
    )


def _campaign_target_bucket_from_values(
    audience_type: str,
    region_level: str,
    region_keys: Optional[List[str]] = None,
    city: Optional[str] = None,
    district: Optional[str] = None,
    neighborhood: Optional[str] = None,
) -> tuple:
    keys = _normalize_region_keys(region_keys) if region_keys is not None else _region_keys_from_legacy_fields(region_level, city, district, neighborhood)
    return (
        _normalize_audience_type(audience_type),
        _normalize_region_level(region_level),
        tuple(sorted(keys)),
    )


def _campaign_target_bucket(campaign: AdCampaign) -> tuple:
    return _campaign_target_bucket_from_values(
        _campaign_audience_type(campaign),
        _campaign_region_level(campaign),
        _campaign_region_keys(campaign),
    )


def _campaign_target_specificity(campaign: AdCampaign) -> tuple:
    region_keys = _campaign_region_keys(campaign)
    return (
        REGION_LEVEL_SCORES.get(_campaign_region_level(campaign), 0),
        1 if _campaign_audience_type(campaign) != 'all' else 0,
        -len(region_keys),
    )


def build_ad_runtime_context(
    current_user: Optional[User] = None,
    *,
    city: Optional[str] = None,
    district: Optional[str] = None,
    neighborhood: Optional[str] = None,
) -> AdRuntimeContext:
    resolved_city = _clean_region_string(city) or _clean_region_string(getattr(current_user, 'last_region_city', None))
    resolved_district = _clean_region_string(district) or _clean_region_string(getattr(current_user, 'last_region_district', None))
    resolved_neighborhood = _clean_region_string(neighborhood) or _clean_region_string(getattr(current_user, 'last_region_neighborhood', None))
    audience_type = 'all'
    if current_user is not None:
        audience_type = 'guest' if bool(getattr(current_user, 'is_guest', False)) else 'member'
    return AdRuntimeContext(
        audience_type=audience_type,
        city=resolved_city,
        district=resolved_district,
        neighborhood=resolved_neighborhood,
    )


def sync_user_region_context(
    db: OrmSession,
    user: Optional[User],
    *,
    city: Optional[str] = None,
    district: Optional[str] = None,
    neighborhood: Optional[str] = None,
    captured_at: Optional[datetime] = None,
    source: Optional[str] = None,
) -> None:
    sync_user_region_context_record(
        db,
        user,
        city=city,
        district=district,
        neighborhood=neighborhood,
        captured_at=captured_at,
        source=source,
    )


def _context_region_key(level: str, context: Optional[AdRuntimeContext]) -> Optional[str]:
    if context is None:
        return None
    if level == 'city':
        return _region_key_from_values('city', context.city)
    if level == 'district':
        return _region_key_from_values('district', context.city, context.district)
    if level == 'neighborhood':
        return _region_key_from_values('neighborhood', context.city, context.district, context.neighborhood)
    return None


def _campaign_matches_runtime_context(campaign: AdCampaign, context: Optional[AdRuntimeContext]) -> bool:
    if context is None:
        return True

    audience_type = _campaign_audience_type(campaign)
    if audience_type != 'all' and audience_type != context.audience_type:
        return False

    region_level = _campaign_region_level(campaign)
    if region_level == 'all':
        return True
    runtime_key = _context_region_key(region_level, context)
    if not runtime_key:
        return False
    return runtime_key in _campaign_region_keys(campaign)


def _region_label_from_key(key: str) -> str:
    parsed = _parse_region_key(key)
    if not parsed:
        return key
    _, city, district, neighborhood = parsed
    return ' / '.join([part for part in [city, district, neighborhood] if part]) or key


def _campaign_target_label(
    audience_type: str,
    region_level: str,
    region_keys: Optional[List[str]] = None,
    city: Optional[str] = None,
    district: Optional[str] = None,
    neighborhood: Optional[str] = None,
) -> str:
    audience_labels = {'all': '전체', 'member': '회원', 'guest': '게스트'}
    region_labels = {'all': '전체지역', 'city': '시', 'district': '구', 'neighborhood': '동'}
    keys = _normalize_region_keys(region_keys) if region_keys is not None else _region_keys_from_legacy_fields(region_level, city, district, neighborhood)
    if region_level == 'all':
        region_detail = '-'
    else:
        labels = [_region_label_from_key(key) for key in keys]
        if not labels:
            region_detail = '-'
        elif len(labels) == 1:
            region_detail = labels[0]
        else:
            region_detail = f'{labels[0]} 외 {len(labels) - 1}'
    return f"{audience_labels.get(audience_type, '전체')} · {region_labels.get(region_level, '전체지역')} · {region_detail}"


def _resolve_primary_region_fields(level: str, region_keys: List[str]) -> tuple[Optional[str], Optional[str], Optional[str]]:
    parsed = [entry for entry in (_parse_region_key(key) for key in region_keys) if entry]
    if not parsed:
        return (None, None, None)
    cities = sorted({entry[1] for entry in parsed if entry[1]})
    districts = sorted({entry[2] for entry in parsed if entry[2]})
    neighborhoods = sorted({entry[3] for entry in parsed if entry[3]})
    city = cities[0] if len(cities) == 1 else None
    district = districts[0] if len(districts) == 1 and level in {'district', 'neighborhood'} else None
    neighborhood = neighborhoods[0] if len(neighborhoods) == 1 and level == 'neighborhood' else None
    return city, district, neighborhood


def _campaign_targets_overlap(
    left_audience: str,
    left_level: str,
    left_keys: List[str],
    right_audience: str,
    right_level: str,
    right_keys: List[str],
) -> bool:
    if left_audience != right_audience or left_level != right_level:
        return False
    if left_level == 'all':
        return True
    return bool(set(left_keys) & set(right_keys))


def _base_slot_config(row: AdSlot, fallback: Dict[str, Any]) -> Dict[str, Any]:
    base_config = fallback['config'].copy()
    if row.config_json:
        try:
            base_config.update(json.loads(row.config_json) or {})
        except Exception:
            pass
    return _normalize_slot_config(base_config, include_reserved=True)


def _campaign_sort_key(campaign: AdCampaign) -> tuple:
    return (
        int(campaign.sort_order or 1),
        campaign.start_at or datetime.min,
        campaign.created_at or datetime.min,
        campaign.id or '',
    )


def _clean_landing_params(value: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(value, dict):
        return None
    cleaned: Dict[str, Any] = {}
    for raw_key, raw_value in value.items():
        key = str(raw_key or '').strip()
        if not key:
            continue
        if isinstance(raw_value, str):
            normalized = raw_value.strip()
            if not normalized:
                continue
            cleaned[key] = normalized
            continue
        if isinstance(raw_value, (int, float, bool)):
            cleaned[key] = raw_value
    return cleaned or None


def _deserialize_landing_params(value: Optional[str]) -> Optional[Dict[str, Any]]:
    if not value or not isinstance(value, str):
        return None
    try:
        parsed = json.loads(value)
    except Exception:
        return None
    return _clean_landing_params(parsed)


def _serialize_campaign_landing(campaign: Optional[AdCampaign]) -> Optional[Dict[str, Any]]:
    if campaign is None:
        return None
    landing_type = str(campaign.landing_type or '').strip()
    landing_key = str(campaign.landing_key or '').strip()
    landing_params = _deserialize_landing_params(campaign.landing_params_json)
    if not landing_type and not landing_key and not landing_params:
        return None
    return {
        'type': landing_type or None,
        'key': landing_key or None,
        'params': landing_params,
    }


def _should_force_full_banner(slot_row: AdSlot) -> bool:
    return str(slot_row.placement_type or '').strip() in {
        'inline',
        'bottom_sheet',
        'floating_overlay',
    }


def _enforce_full_banner_payload(slot_row: AdSlot, payload: Dict[str, Any]) -> Dict[str, Any]:
    if not _should_force_full_banner(slot_row):
        return payload
    landing_params = dict(payload.get('landingParams') or {})
    landing_params['renderStyle'] = 'full_banner'
    payload['landingParams'] = _clean_landing_params(landing_params)
    return payload


def _campaign_window_key(campaign: AdCampaign) -> tuple:
    return (
        campaign.start_at.isoformat() if campaign.start_at else '',
        campaign.end_at.isoformat() if campaign.end_at else '',
    )


def _runtime_group_sort_key(group: List[AdCampaign]) -> tuple:
    lead = sorted(group, key=_campaign_sort_key)[0]
    return (
        lead.start_at or datetime.max,
        lead.end_at or datetime.max,
        int(lead.sort_order or 1),
        lead.created_at or datetime.max,
        lead.id or '',
    )


def _group_campaigns_by_window(campaigns: List[AdCampaign]) -> List[List[AdCampaign]]:
    grouped: Dict[tuple, List[AdCampaign]] = {}
    for campaign in campaigns:
        grouped.setdefault(_campaign_window_key(campaign), []).append(campaign)
    groups = [sorted(group, key=_campaign_sort_key) for group in grouped.values()]
    groups.sort(key=_runtime_group_sort_key)
    return groups


def _campaign_group_rotation_mode(group: List[AdCampaign]) -> str:
    if len(group) <= 1:
        return 'single'
    if all(int(campaign.sort_order or 1) == 999 for campaign in group):
        return 'random'
    return 'ordered'


def _primary_campaign_from_group(group: List[AdCampaign]) -> Optional[AdCampaign]:
    return sorted(group, key=_campaign_sort_key)[0] if group else None


def _serialize_runtime_creative(campaign: AdCampaign) -> Dict[str, Any]:
    landing = _serialize_campaign_landing(campaign) or {}
    audience_type = _campaign_audience_type(campaign)
    region_level = _campaign_region_level(campaign)
    region_keys = _campaign_region_keys(campaign)
    target_city = _clean_region_string(campaign.target_city)
    target_district = _clean_region_string(campaign.target_district)
    target_neighborhood = _clean_region_string(campaign.target_neighborhood)
    return {
        'campaignId': campaign.id,
        'creativeId': campaign.id,
        'title': campaign.title or '',
        'message': campaign.message or '',
        'ctaLabel': campaign.cta_label,
        'targetUrl': campaign.target_url,
        'imageUrl': campaign.image_url,
        'sortOrder': int(campaign.sort_order or 1),
        'startAt': campaign.start_at.isoformat() if campaign.start_at else None,
        'endAt': campaign.end_at.isoformat() if campaign.end_at else None,
        'landingType': landing.get('type'),
        'landingKey': landing.get('key'),
        'landingParams': landing.get('params'),
        'audienceType': audience_type,
        'targetRegionLevel': region_level,
        'targetCity': target_city,
        'targetDistrict': target_district,
        'targetNeighborhood': target_neighborhood,
        'targetRegionKeys': region_keys,
        'targetLabel': _campaign_target_label(audience_type, region_level, region_keys, target_city, target_district, target_neighborhood),
    }


def _best_matching_campaign_groups(campaigns: List[AdCampaign], target_context: Optional[AdRuntimeContext]) -> List[List[AdCampaign]]:
    if target_context is None:
        return _group_campaigns_by_window(campaigns)
    matching_campaigns = [campaign for campaign in campaigns if _campaign_matches_runtime_context(campaign, target_context)]
    if not matching_campaigns:
        return []
    best_campaign = max(matching_campaigns, key=_campaign_target_specificity)
    best_bucket = _campaign_target_bucket(best_campaign)
    return _group_campaigns_by_window([campaign for campaign in matching_campaigns if _campaign_target_bucket(campaign) == best_bucket])


def _derive_runtime_groups(campaigns: List[AdCampaign], target_context: Optional[AdRuntimeContext] = None) -> tuple[List[AdCampaign], List[AdCampaign]]:
    active_campaigns = [campaign for campaign in campaigns if _campaign_runtime_status(campaign) != 'cancelled']
    live_candidates = [campaign for campaign in active_campaigns if _campaign_runtime_status(campaign) == 'live']
    scheduled_candidates = [campaign for campaign in active_campaigns if _campaign_runtime_status(campaign) == 'scheduled']

    live_groups = _best_matching_campaign_groups(live_candidates, target_context)
    scheduled_groups = _best_matching_campaign_groups(scheduled_candidates, target_context)
    return (live_groups[0] if live_groups else []), (scheduled_groups[0] if scheduled_groups else [])


def _derive_runtime_campaigns(campaigns: List[AdCampaign], target_context: Optional[AdRuntimeContext] = None) -> tuple[Optional[AdCampaign], Optional[AdCampaign]]:
    live_group, next_group = _derive_runtime_groups(campaigns, target_context=target_context)
    return _primary_campaign_from_group(live_group), _primary_campaign_from_group(next_group)


def _campaign_fields(campaign: Optional[AdCampaign], *, include_reserved: bool = False) -> Dict[str, Any]:
    if campaign is None:
        if include_reserved:
            return {
                'reservedTitle': '',
                'reservedMessage': '',
                'reservedCtaLabel': None,
                'reservedTargetUrl': None,
                'reservedLandingType': None,
                'reservedLandingKey': None,
                'reservedLandingParams': None,
                'reservedImageUrl': None,
                'reservationStartAt': None,
                'reservationEndAt': None,
                'reservedCampaignId': None,
            }
        return {
            'title': '',
            'message': '',
            'ctaLabel': None,
            'targetUrl': None,
            'landingType': None,
            'landingKey': None,
            'landingParams': None,
            'imageUrl': None,
            'exposureStartAt': None,
            'exposureEndAt': None,
            'liveCampaignId': None,
        }

    landing = _serialize_campaign_landing(campaign) or {}
    if include_reserved:
        return {
            'reservedTitle': campaign.title or '',
            'reservedMessage': campaign.message or '',
            'reservedCtaLabel': campaign.cta_label,
            'reservedTargetUrl': campaign.target_url,
            'reservedLandingType': landing.get('type'),
            'reservedLandingKey': landing.get('key'),
            'reservedLandingParams': landing.get('params'),
            'reservedImageUrl': campaign.image_url,
            'reservationStartAt': campaign.start_at.isoformat() if campaign.start_at else None,
            'reservationEndAt': campaign.end_at.isoformat() if campaign.end_at else None,
            'reservedCampaignId': campaign.id,
        }
    return {
        'title': campaign.title or '',
        'message': campaign.message or '',
        'ctaLabel': campaign.cta_label,
        'targetUrl': campaign.target_url,
        'landingType': landing.get('type'),
        'landingKey': landing.get('key'),
        'landingParams': landing.get('params'),
        'imageUrl': campaign.image_url,
        'exposureStartAt': campaign.start_at.isoformat() if campaign.start_at else None,
        'exposureEndAt': campaign.end_at.isoformat() if campaign.end_at else None,
        'liveCampaignId': campaign.id,
    }


def _project_slot_config_from_campaigns(
    base_config: Dict[str, Any],
    campaigns: List[AdCampaign],
    *,
    target_context: Optional[AdRuntimeContext] = None,
) -> Dict[str, Any]:
    projected = base_config.copy()
    live_group, next_group = _derive_runtime_groups(campaigns, target_context=target_context)
    live_campaign = _primary_campaign_from_group(live_group)
    next_campaign = _primary_campaign_from_group(next_group)
    projected.update(_campaign_fields(live_campaign))
    projected.update(_campaign_fields(next_campaign, include_reserved=True))
    projected['liveCreatives'] = [_serialize_runtime_creative(campaign) for campaign in live_group]
    projected['liveRotationMode'] = _campaign_group_rotation_mode(live_group)
    projected['reservedCreatives'] = [_serialize_runtime_creative(campaign) for campaign in next_group]
    projected['reservedRotationMode'] = _campaign_group_rotation_mode(next_group)
    return _normalize_slot_config(projected, include_reserved=True)


def _sync_slot_runtime_config(db: OrmSession, row: AdSlot, fallback: Dict[str, Any]) -> None:
    _ensure_campaign_links_for_row(db, row, fallback)
    campaigns = db.scalars(select(AdCampaign).where(AdCampaign.slot_id == row.id).order_by(AdCampaign.created_at.asc())).all()
    next_config = _project_slot_config_from_campaigns(_base_slot_config(row, fallback), campaigns)
    current_config = _base_slot_config(row, fallback)
    if current_config != next_config:
        row.config_json = json.dumps(next_config, ensure_ascii=False)
        db.add(row)


def _normalize_campaign_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    def _clean_optional(value: Any) -> Optional[str]:
        if value is None:
            return None
        if not isinstance(value, str):
            value = str(value)
        cleaned = value.strip()
        return cleaned or None

    sort_order_raw = payload.get('sortOrder')
    try:
        sort_order = int(sort_order_raw) if sort_order_raw is not None else 1
    except (TypeError, ValueError):
        sort_order = 1
    if sort_order <= 0:
        sort_order = 1

    title = _clean_optional(payload.get('title')) or ''
    message = _clean_optional(payload.get('message')) or ''
    target_city = _clean_region_string(payload.get('targetCity'))
    target_district = _clean_region_string(payload.get('targetDistrict'))
    target_neighborhood = _clean_region_string(payload.get('targetNeighborhood'))
    target_region_keys = _normalize_region_keys(payload.get('targetRegionKeys'))
    target_region_level = _normalize_region_level(payload.get('targetRegionLevel'))
    if target_region_level == 'all' and target_region_keys:
        parsed = _parse_region_key(target_region_keys[0])
        if parsed:
            target_region_level = parsed[0]
    if target_region_level == 'all':
        target_region_level = _resolve_target_region_level(
            payload.get('targetRegionLevel'),
            target_city,
            target_district,
            target_neighborhood,
        )
    if not target_region_keys and target_region_level != 'all':
        target_region_keys = _region_keys_from_legacy_fields(target_region_level, target_city, target_district, target_neighborhood)
    return {
        'slotKey': _clean_optional(payload.get('slotKey')),
        'sortOrder': sort_order,
        'title': title,
        'message': message,
        'ctaLabel': _clean_optional(payload.get('ctaLabel')),
        'targetUrl': _clean_optional(payload.get('targetUrl')),
        'landingType': _clean_optional(payload.get('landingType')),
        'landingKey': _clean_optional(payload.get('landingKey')),
        'landingParams': _clean_landing_params(payload.get('landingParams')),
        'imageUrl': _clean_optional(payload.get('imageUrl')),
        'audienceType': _normalize_audience_type(payload.get('audienceType')),
        'targetRegionLevel': target_region_level,
        'targetCity': target_city,
        'targetDistrict': target_district,
        'targetNeighborhood': target_neighborhood,
        'targetRegionKeys': target_region_keys,
        'startAt': _clean_optional(payload.get('startAt')),
        'endAt': _clean_optional(payload.get('endAt')),
    }


def _campaign_overlaps(start_at: Optional[datetime], end_at: Optional[datetime], other_start_at: Optional[datetime], other_end_at: Optional[datetime]) -> bool:
    left_start = start_at or datetime.min
    left_end = end_at or datetime.max
    right_start = other_start_at or datetime.min
    right_end = other_end_at or datetime.max
    return left_start <= right_end and right_start <= left_end


def _campaign_identity_signature(payload: Dict[str, Any]) -> tuple:
    landing_params = payload.get('landingParams') or {}
    return (
        payload.get('slotKey'),
        payload.get('startAt'),
        payload.get('endAt'),
        payload.get('title') or '',
        payload.get('message') or '',
        payload.get('ctaLabel') or '',
        payload.get('targetUrl') or '',
        payload.get('landingType') or '',
        payload.get('landingKey') or '',
        json.dumps(landing_params, ensure_ascii=False, sort_keys=True),
        payload.get('imageUrl') or '',
        payload.get('audienceType') or 'all',
        payload.get('targetRegionLevel') or 'all',
        json.dumps(sorted(payload.get('targetRegionKeys') or []), ensure_ascii=False),
    )


def _validate_campaign_payload(db: OrmSession, slot_row: AdSlot, payload: Dict[str, Any], *, exclude_campaign_id: Optional[str] = None) -> Dict[str, Any]:
    normalized = _enforce_full_banner_payload(slot_row, _normalize_campaign_payload(payload))
    if not normalized.get('slotKey'):
        raise ValueError('slotKey is required')
    if not normalized.get('startAt') or not normalized.get('endAt'):
        raise ValueError('startAt and endAt are required')

    start_at = _parse_optional_datetime(normalized.get('startAt'))
    end_at = _parse_optional_datetime(normalized.get('endAt'))
    if start_at is None or end_at is None:
        raise ValueError('startAt/endAt must be valid ISO datetime values')
    if end_at <= start_at:
        raise ValueError('endAt must be later than startAt')
    if not any([normalized.get('title'), normalized.get('message'), normalized.get('imageUrl')]):
        raise ValueError('title, message, image 중 하나는 채워야 해')

    landing_type = normalized.get('landingType')
    landing_key = normalized.get('landingKey')
    allowed_landing_types = {'explore_section', 'my_section', 'auth_flow', 'saved_flow', 'home_tab'}
    if landing_type and landing_type not in allowed_landing_types:
        raise ValueError('지원하지 않는 landingType 이야')
    if landing_type and not landing_key:
        raise ValueError('landingKey 가 필요해')
    if landing_key and not landing_type:
        raise ValueError('landingType 이 필요해')
    if normalized.get('targetUrl') and landing_type:
        raise ValueError('링크URL 과 랜딩페이지 중 하나만 써줘')

    audience_type = _normalize_audience_type(normalized.get('audienceType'))
    if audience_type not in AUDIENCE_TYPES:
        raise ValueError('고객구분은 전체, 회원, 게스트 중 하나여야 해')

    target_region_level = _normalize_region_level(normalized.get('targetRegionLevel'))
    target_region_keys = _normalize_region_keys(normalized.get('targetRegionKeys'))
    if target_region_level not in REGION_LEVELS:
        raise ValueError('지역단위는 전체, 시, 구, 동 중 하나여야 해')
    if target_region_level == 'all':
        target_region_keys = []
    else:
        if not target_region_keys:
            raise ValueError('지역 타겟팅에는 최소 1개 이상 선택이 필요해')
        key_levels = {(_parse_region_key(key) or ('all', None, None, None))[0] for key in target_region_keys}
        if len(key_levels) != 1 or target_region_level not in key_levels:
            raise ValueError('선택한 지역키와 지역단위가 맞지 않아')

    target_city, target_district, target_neighborhood = _resolve_primary_region_fields(target_region_level, target_region_keys)

    normalized['audienceType'] = audience_type
    normalized['targetRegionLevel'] = target_region_level
    normalized['targetRegionKeys'] = target_region_keys
    normalized['targetCity'] = target_city
    normalized['targetDistrict'] = target_district
    normalized['targetNeighborhood'] = target_neighborhood

    existing_campaigns = db.scalars(select(AdCampaign).where(AdCampaign.slot_id == slot_row.id)).all()
    target_signature = _campaign_identity_signature(normalized)
    for campaign in existing_campaigns:
        if exclude_campaign_id and campaign.id == exclude_campaign_id:
            continue
        if _campaign_runtime_status(campaign) == 'cancelled':
            continue
        existing_keys = _campaign_region_keys(campaign)
        existing_payload = {
            'slotKey': normalized['slotKey'],
            'startAt': campaign.start_at.isoformat() if campaign.start_at else None,
            'endAt': campaign.end_at.isoformat() if campaign.end_at else None,
            'title': campaign.title or '',
            'message': campaign.message or '',
            'ctaLabel': campaign.cta_label or '',
            'targetUrl': campaign.target_url or '',
            'landingType': campaign.landing_type or '',
            'landingKey': campaign.landing_key or '',
            'landingParams': _deserialize_landing_params(campaign.landing_params_json) or {},
            'imageUrl': campaign.image_url or '',
            'audienceType': _campaign_audience_type(campaign),
            'targetRegionLevel': _campaign_region_level(campaign),
            'targetRegionKeys': existing_keys,
        }
        if target_signature == _campaign_identity_signature(existing_payload):
            raise ValueError('같은 슬롯, 시간, 타겟, 소재 row가 이미 있어')

        if not _campaign_targets_overlap(
            audience_type,
            target_region_level,
            target_region_keys,
            _campaign_audience_type(campaign),
            _campaign_region_level(campaign),
            existing_keys,
        ):
            continue

        exact_same_window = campaign.start_at == start_at and campaign.end_at == end_at
        if exact_same_window:
            existing_sort = int(campaign.sort_order or 1)
            next_sort = int(normalized.get('sortOrder') or 1)
            if (existing_sort == 999) != (next_sort == 999):
                raise ValueError('같은 타겟 기간 묶음은 전부 순차 정렬이거나 전부 999 랜덤이어야 해')
            if next_sort != 999 and existing_sort != 999 and next_sort == existing_sort:
                raise ValueError(f'같은 타겟 기간 묶음 안에서 정렬 {next_sort} 이 중복돼')
        if _campaign_overlaps(start_at, end_at, campaign.start_at, campaign.end_at) and not exact_same_window:
            label = campaign.title or campaign.message or campaign.id
            existing_start = campaign.start_at.isoformat() if campaign.start_at else '-'
            existing_end = campaign.end_at.isoformat() if campaign.end_at else '-'
            raise ValueError(f'같은 타겟 슬롯 시간대가 일부 겹쳐. 동일 기간 묶음만 함께 둘 수 있어. existing={label} ({existing_start} ~ {existing_end})')

    normalized['startAt'] = start_at
    normalized['endAt'] = end_at
    return normalized


def list_ad_slots(db: OrmSession) -> List[Dict[str, Any]]:
    ensure_default_ad_slots(db)
    rows = db.scalars(select(AdSlot).order_by(AdSlot.created_at.asc())).all()
    defaults = _defaults_by_key()
    dirty = False
    for row in rows:
        if row.slot_key in defaults:
            before = row.config_json
            _sync_slot_runtime_config(db, row, defaults[row.slot_key])
            if row.config_json != before:
                dirty = True
    if dirty:
        db.commit()
        for row in rows:
            db.refresh(row)
    return [_serialize_slot(row, defaults[row.slot_key]) for row in rows if row.slot_key in defaults]


def update_ad_slot(db: OrmSession, slot_key: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    ensure_default_ad_slots(db)
    row = db.scalar(select(AdSlot).where(AdSlot.slot_key == slot_key))
    if row is None:
        raise ValueError('slot_not_found')

    defaults = _defaults_by_key()
    base_config = _base_slot_config(row, defaults[slot_key])
    next_config = base_config.copy()
    for key in ['slotLabel', 'slotDescription', 'placementNote']:
        if key in (payload or {}):
            next_config[key] = payload.get(key)

    if 'status' in (payload or {}):
        row.status = str((payload or {}).get('status') or row.status or 'active').strip() or 'active'

    row.config_json = json.dumps(_normalize_slot_config(next_config, include_reserved=True), ensure_ascii=False)
    db.add(row)
    _sync_slot_runtime_config(db, row, defaults[slot_key])
    db.commit()
    db.refresh(row)
    return _serialize_slot(row, defaults[slot_key])


def create_ad_campaign(db: OrmSession, payload: Dict[str, Any]) -> Dict[str, Any]:
    ensure_default_ad_slots(db)
    normalized = _normalize_campaign_payload(payload)
    slot_key = normalized.get('slotKey')
    slot_row = db.scalar(select(AdSlot).where(AdSlot.slot_key == slot_key)) if slot_key else None
    if slot_row is None:
        raise ValueError('slot_not_found')

    validated = _validate_campaign_payload(db, slot_row, normalized)
    now = datetime.now()
    campaign = AdCampaign(
        slot_id=slot_row.id,
        variant='reserved' if validated['startAt'] > now else 'live',
        status=_campaign_runtime_status(
            AdCampaign(start_at=validated['startAt'], end_at=validated['endAt']),
            now=now,
        ),
        title=validated['title'] or '',
        message=validated['message'] or '',
        cta_label=validated['ctaLabel'],
        target_url=validated['targetUrl'],
        landing_type=validated['landingType'],
        landing_key=validated['landingKey'],
        landing_params_json=json.dumps(validated['landingParams'], ensure_ascii=False) if validated.get('landingParams') else None,
        image_url=validated['imageUrl'],
        audience_type=validated['audienceType'],
        target_region_level=validated['targetRegionLevel'],
        target_city=validated['targetCity'],
        target_district=validated['targetDistrict'],
        target_neighborhood=validated['targetNeighborhood'],
        target_region_keys_json=json.dumps(validated['targetRegionKeys'], ensure_ascii=False) if validated.get('targetRegionKeys') else None,
        sort_order=int(validated.get('sortOrder') or 1),
        start_at=validated['startAt'],
        end_at=validated['endAt'],
    )
    db.add(campaign)
    db.flush()
    defaults = _defaults_by_key()
    _sync_slot_runtime_config(db, slot_row, defaults[slot_key])
    db.commit()
    db.refresh(campaign)
    return _serialize_campaign_metrics(db, [campaign], {slot_row.id: slot_key})[0]


def update_ad_campaign(db: OrmSession, campaign_id: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    ensure_default_ad_slots(db)
    campaign = db.get(AdCampaign, campaign_id)
    if campaign is None:
        raise ValueError('campaign_not_found')

    normalized = _normalize_campaign_payload(payload)
    slot_key = normalized.get('slotKey')
    target_slot_row = db.scalar(select(AdSlot).where(AdSlot.slot_key == slot_key)) if slot_key else None
    if target_slot_row is None:
        raise ValueError('slot_not_found')

    previous_slot_row = db.get(AdSlot, campaign.slot_id)
    validated = _validate_campaign_payload(db, target_slot_row, normalized, exclude_campaign_id=campaign.id)
    now = datetime.now()

    campaign.slot_id = target_slot_row.id
    campaign.title = validated['title'] or ''
    campaign.message = validated['message'] or ''
    campaign.cta_label = validated['ctaLabel']
    campaign.target_url = validated['targetUrl']
    campaign.landing_type = validated['landingType']
    campaign.landing_key = validated['landingKey']
    campaign.landing_params_json = json.dumps(validated['landingParams'], ensure_ascii=False) if validated.get('landingParams') else None
    campaign.image_url = validated['imageUrl']
    campaign.audience_type = validated['audienceType']
    campaign.target_region_level = validated['targetRegionLevel']
    campaign.target_city = validated['targetCity']
    campaign.target_district = validated['targetDistrict']
    campaign.target_neighborhood = validated['targetNeighborhood']
    campaign.target_region_keys_json = json.dumps(validated['targetRegionKeys'], ensure_ascii=False) if validated.get('targetRegionKeys') else None
    campaign.sort_order = int(validated.get('sortOrder') or 1)
    campaign.start_at = validated['startAt']
    campaign.end_at = validated['endAt']
    campaign.variant = 'reserved' if validated['startAt'] > now else 'live'
    campaign.status = _campaign_runtime_status(campaign, now=now)
    db.add(campaign)

    defaults = _defaults_by_key()
    if previous_slot_row is not None and previous_slot_row.slot_key in defaults:
        _sync_slot_runtime_config(db, previous_slot_row, defaults[previous_slot_row.slot_key])
    _sync_slot_runtime_config(db, target_slot_row, defaults[target_slot_row.slot_key])
    db.commit()
    db.refresh(campaign)
    return _serialize_campaign_metrics(db, [campaign], {target_slot_row.id: target_slot_row.slot_key})[0]


def cancel_ad_campaign(db: OrmSession, campaign_id: str) -> Dict[str, Any]:
    ensure_default_ad_slots(db)
    campaign = db.get(AdCampaign, campaign_id)
    if campaign is None:
        raise ValueError('campaign_not_found')

    campaign.status = 'cancelled'
    db.add(campaign)
    slot_row = db.get(AdSlot, campaign.slot_id)
    defaults = _defaults_by_key()
    if slot_row is not None and slot_row.slot_key in defaults:
        _sync_slot_runtime_config(db, slot_row, defaults[slot_row.slot_key])
    db.commit()
    db.refresh(campaign)
    if slot_row is not None:
        return _serialize_campaign_metrics(db, [campaign], {slot_row.id: slot_row.slot_key})[0]
    return {'id': campaign.id, 'status': 'cancelled'}


def delete_ad_campaign(db: OrmSession, campaign_id: str) -> Dict[str, Any]:
    ensure_default_ad_slots(db)
    campaign = db.get(AdCampaign, campaign_id)
    if campaign is None:
        raise ValueError('campaign_not_found')

    slot_row = db.get(AdSlot, campaign.slot_id)
    result = {'id': campaign.id, 'slotKey': slot_row.slot_key if slot_row is not None else None, 'deleted': True}
    db.delete(campaign)
    defaults = _defaults_by_key()
    if slot_row is not None and slot_row.slot_key in defaults:
        _sync_slot_runtime_config(db, slot_row, defaults[slot_row.slot_key])
    db.commit()
    return result


def _parse_period_start(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(f'{value.strip()}T00:00:00')
    except ValueError:
        return None


def _parse_period_end(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(f'{value.strip()}T23:59:59')
    except ValueError:
        return None


def _slot_surface_label(slot: Dict[str, Any]) -> str:
    config = slot.get('config') or {}
    screen = str(config.get('screen') or '').strip()
    position = str(config.get('position') or '').strip()
    parts = [part for part in [screen, position] if part]
    return ' · '.join(parts) if parts else slot.get('slotKey') or '-'


def _slot_placement_label(slot: Dict[str, Any]) -> str:
    config = slot.get('config') or {}
    placement_note = str(config.get('placementNote') or '').strip()
    if placement_note:
        return placement_note
    return str(slot.get('placementType') or '-').strip() or '-'


def _effective_runtime_state(slot: Dict[str, Any]) -> str:
    if slot.get('status') != 'active':
        return 'inactive'

    config = slot.get('config') or {}
    now = datetime.now()
    live_title = str(config.get('title') or '').strip()
    reserved_title = str(config.get('reservedTitle') or '').strip()
    live_start = config.get('exposureStartAt')
    live_end = config.get('exposureEndAt')
    reserved_start = config.get('reservationStartAt')
    reserved_end = config.get('reservationEndAt')

    live_start_dt = datetime.fromisoformat(live_start) if isinstance(live_start, str) and live_start.strip() else None
    live_end_dt = datetime.fromisoformat(live_end) if isinstance(live_end, str) and live_end.strip() else None
    reserved_start_dt = datetime.fromisoformat(reserved_start) if isinstance(reserved_start, str) and reserved_start.strip() else None

    live_window_ok = (live_start_dt is None or live_start_dt <= now) and (live_end_dt is None or live_end_dt >= now)
    reserved_future = reserved_start_dt is not None and reserved_start_dt > now

    if live_title and live_window_ok:
        return 'live_now'
    if live_title and live_end_dt is not None and live_end_dt < now:
        return 'expired'
    if reserved_title and reserved_future:
        return 'reserved_pending'
    if reserved_title and not reserved_start_dt:
        return 'reserved_mismatch'
    if not live_title and reserved_title:
        return 'draft_gap'
    if not live_title:
        return 'inactive_gap'
    return 'draft_gap'


def _serialize_campaign_metrics(
    db: OrmSession,
    campaigns: List[AdCampaign],
    slot_key_map: Dict[str, str],
    *,
    period_start: Optional[datetime] = None,
    period_end: Optional[datetime] = None,
) -> List[Dict[str, Any]]:
    if not campaigns:
        return []

    campaign_ids = [campaign.id for campaign in campaigns]

    impression_query = select(AdImpression.campaign_id, func.count(AdImpression.id)).where(AdImpression.campaign_id.in_(campaign_ids))
    click_query = (
        select(AdImpression.campaign_id, func.count(AdClick.id))
        .join(AdClick, AdClick.impression_id == AdImpression.id)
        .where(AdImpression.campaign_id.in_(campaign_ids))
    )
    if period_start is not None:
        impression_query = impression_query.where(AdImpression.created_at >= period_start)
        click_query = click_query.where(AdImpression.created_at >= period_start)
    if period_end is not None:
        impression_query = impression_query.where(AdImpression.created_at <= period_end)
        click_query = click_query.where(AdImpression.created_at <= period_end)

    impression_rows = db.execute(impression_query.group_by(AdImpression.campaign_id)).all()
    click_rows = db.execute(click_query.group_by(AdImpression.campaign_id)).all()
    impression_map = {campaign_id: count for campaign_id, count in impression_rows}
    click_map = {campaign_id: count for campaign_id, count in click_rows}

    result = []
    for campaign in campaigns:
        impressions = int(impression_map.get(campaign.id, 0) or 0)
        clicks = int(click_map.get(campaign.id, 0) or 0)
        ctr = 0.0 if impressions == 0 else round(clicks / impressions, 4)
        landing = _serialize_campaign_landing(campaign) or {}
        audience_type = _campaign_audience_type(campaign)
        region_level = _campaign_region_level(campaign)
        region_keys = _campaign_region_keys(campaign)
        target_city = _clean_region_string(campaign.target_city)
        target_district = _clean_region_string(campaign.target_district)
        target_neighborhood = _clean_region_string(campaign.target_neighborhood)
        result.append(
            {
                'id': campaign.id,
                'slotKey': slot_key_map.get(campaign.slot_id),
                'variant': _campaign_runtime_variant(campaign),
                'status': _campaign_runtime_status(campaign),
                'title': campaign.title,
                'message': campaign.message,
                'ctaLabel': campaign.cta_label,
                'targetUrl': campaign.target_url,
                'landingType': landing.get('type'),
                'landingKey': landing.get('key'),
                'landingParams': landing.get('params'),
                'imageUrl': campaign.image_url,
                'audienceType': audience_type,
                'targetRegionLevel': region_level,
                'targetCity': target_city,
                'targetDistrict': target_district,
                'targetNeighborhood': target_neighborhood,
                'targetRegionKeys': region_keys,
                'targetLabel': _campaign_target_label(audience_type, region_level, region_keys, target_city, target_district, target_neighborhood),
                'sortOrder': int(campaign.sort_order or 1),
                'startAt': campaign.start_at.isoformat() if campaign.start_at else None,
                'endAt': campaign.end_at.isoformat() if campaign.end_at else None,
                'impressions': impressions,
                'clicks': clicks,
                'ctr': ctr,
                'createdAt': campaign.created_at.isoformat() if campaign.created_at else None,
                'updatedAt': campaign.updated_at.isoformat() if campaign.updated_at else None,
            }
        )
    return result


def list_ad_campaigns(
    db: OrmSession,
    slot_key: Optional[str] = None,
    limit: int = 100,
    *,
    query: Optional[str] = None,
    variant: Optional[str] = None,
    status: Optional[str] = None,
    period_from: Optional[str] = None,
    period_to: Optional[str] = None,
) -> List[Dict[str, Any]]:
    ensure_default_ad_slots(db)

    slot_query = select(AdSlot)
    if slot_key:
        slot_query = slot_query.where(AdSlot.slot_key == slot_key)
    slots = db.scalars(slot_query).all()
    if not slots:
        return []

    slot_ids = [slot.id for slot in slots]
    slot_key_map = {slot.id: slot.slot_key for slot in slots}

    campaign_query = select(AdCampaign).where(AdCampaign.slot_id.in_(slot_ids)).order_by(AdCampaign.created_at.desc())
    campaigns = [
        campaign
        for campaign in db.scalars(campaign_query).all()
        if _campaign_matches_filters(
            campaign,
            query=query,
            variant=variant,
            status=status,
            period_from=period_from,
            period_to=period_to,
        )
    ][:limit]
    return _serialize_campaign_metrics(db, campaigns, slot_key_map)


def get_ad_performance_summary(
    db: OrmSession,
    *,
    slot_key: Optional[str] = None,
    surface: Optional[str] = None,
    status: Optional[str] = None,
    variant: Optional[str] = None,
    period_from: Optional[str] = None,
    period_to: Optional[str] = None,
) -> Dict[str, Any]:
    slots = list_ad_slots(db)
    empty = {
        'summary': {'liveSlots': 0, 'reservedPending': 0, 'activeCreatives': 0, 'lowCtrSlots': 0, 'noDataSlots': 0},
        'slotRows': [],
        'creativeRows': [],
        'reviewQueues': {'noData': [], 'lowCtr': [], 'inactiveGap': [], 'reservedMismatch': [], 'clickNoDownstream': []},
    }
    if not slots:
        return empty

    filtered_slots = []
    for slot in slots:
        if slot_key and slot.get('slotKey') != slot_key:
            continue
        if status and status not in {'', 'all'} and status in {'active', 'inactive'} and slot.get('status') != status:
            continue
        surface_haystack = ' '.join(
            str(part or '')
            for part in [
                slot.get('slotKey'),
                (slot.get('config') or {}).get('screen'),
                (slot.get('config') or {}).get('position'),
                (slot.get('config') or {}).get('placementNote'),
            ]
        ).lower()
        if surface and surface.strip() and surface.strip().lower() not in surface_haystack:
            continue
        filtered_slots.append(slot)

    if not filtered_slots:
        return empty

    slot_ids = [slot['id'] for slot in filtered_slots]
    slot_key_map = {slot['id']: slot['slotKey'] for slot in filtered_slots}
    period_start = _parse_period_start(period_from)
    period_end = _parse_period_end(period_to)

    campaign_query = select(AdCampaign).where(AdCampaign.slot_id.in_(slot_ids)).order_by(AdCampaign.created_at.desc())
    campaigns = [
        campaign
        for campaign in db.scalars(campaign_query).all()
        if _campaign_matches_filters(
            campaign,
            variant=variant,
            status=status if status not in {None, '', 'all', 'active', 'inactive'} else None,
            period_from=period_from,
            period_to=period_to,
        )
    ]
    campaign_rows = _serialize_campaign_metrics(db, campaigns, slot_key_map, period_start=period_start, period_end=period_end)
    campaign_by_id = {row['id']: row for row in campaign_rows}

    impression_query = select(AdImpression.slot_id, func.count(AdImpression.id), func.max(AdImpression.created_at)).where(AdImpression.slot_id.in_(slot_ids))
    click_query = (
        select(AdImpression.slot_id, func.count(AdClick.id))
        .join(AdClick, AdClick.impression_id == AdImpression.id)
        .where(AdImpression.slot_id.in_(slot_ids))
    )
    if campaign_by_id:
        campaign_ids = list(campaign_by_id.keys())
        impression_query = impression_query.where(AdImpression.campaign_id.in_(campaign_ids))
        click_query = click_query.where(AdImpression.campaign_id.in_(campaign_ids))
    if period_start is not None:
        impression_query = impression_query.where(AdImpression.created_at >= period_start)
        click_query = click_query.where(AdImpression.created_at >= period_start)
    if period_end is not None:
        impression_query = impression_query.where(AdImpression.created_at <= period_end)
        click_query = click_query.where(AdImpression.created_at <= period_end)

    slot_impression_rows = db.execute(impression_query.group_by(AdImpression.slot_id)).all()
    slot_click_rows = db.execute(click_query.group_by(AdImpression.slot_id)).all()
    slot_impressions = {slot_id: int(count or 0) for slot_id, count, _ in slot_impression_rows}
    slot_last_impression = {slot_id: latest.isoformat() if latest else None for slot_id, _, latest in slot_impression_rows}
    slot_clicks = {slot_id: int(count or 0) for slot_id, count in slot_click_rows}

    creative_rows = [
        {
            'creativeKey': row['id'],
            'title': row['title'] or '(제목 없음)',
            'slotCount': 1,
            'slotKey': row['slotKey'],
            'variant': row['variant'],
            'status': row['status'],
            'impressions': row['impressions'],
            'clicks': row['clicks'],
            'ctr': row['ctr'],
            'downstreamActions': None,
            'firstSeenAt': row['startAt'] or row['createdAt'],
            'lastSeenAt': row['updatedAt'] or row['endAt'],
        }
        for row in campaign_rows
    ]

    slot_rows = []
    review_queues = {'noData': [], 'lowCtr': [], 'inactiveGap': [], 'reservedMismatch': [], 'clickNoDownstream': []}
    for slot in filtered_slots:
        config = slot.get('config') or {}
        live_campaign = campaign_by_id.get(config.get('liveCampaignId'))
        reserved_campaign = campaign_by_id.get(config.get('reservedCampaignId'))
        impressions = slot_impressions.get(slot['id'], 0)
        clicks = slot_clicks.get(slot['id'], 0)
        ctr = 0.0 if impressions == 0 else round(clicks / impressions, 4)
        runtime_state = _effective_runtime_state(slot)
        review_flag = 'ok'
        if runtime_state == 'reserved_mismatch':
            review_flag = 'reserved_mismatch'
        elif runtime_state == 'inactive_gap':
            review_flag = 'inactive_gap'
        elif runtime_state in {'live_now', 'reserved_pending', 'draft_gap'} and impressions == 0:
            review_flag = 'no_data'
        elif impressions >= 20 and ctr < 0.02:
            review_flag = 'low_ctr'

        slot_row = {
            'slotKey': slot['slotKey'],
            'surfaceLabel': _slot_surface_label(slot),
            'placementLabel': _slot_placement_label(slot),
            'slotStatus': slot['status'],
            'effectiveRuntimeState': runtime_state,
            'liveCreativeTitle': (live_campaign or {}).get('title') or config.get('title') or '-',
            'livePeriod': {'startAt': config.get('exposureStartAt'), 'endAt': config.get('exposureEndAt')},
            'reservedCreativeTitle': (reserved_campaign or {}).get('title') or config.get('reservedTitle') or '-',
            'reservedPeriod': {'startAt': config.get('reservationStartAt'), 'endAt': config.get('reservationEndAt')},
            'updatedAt': slot.get('updatedAt') or slot.get('createdAt'),
            'impressions': impressions,
            'clicks': clicks,
            'ctr': ctr,
            'downstreamActions': None,
            'reviewFlag': review_flag,
            'lastImpressionAt': slot_last_impression.get(slot['id']),
        }
        slot_rows.append(slot_row)

        if review_flag == 'no_data':
            review_queues['noData'].append(slot_row)
        elif review_flag == 'low_ctr':
            review_queues['lowCtr'].append(slot_row)
        elif review_flag == 'inactive_gap':
            review_queues['inactiveGap'].append(slot_row)
        elif review_flag == 'reserved_mismatch':
            review_queues['reservedMismatch'].append(slot_row)

    summary = {
        'liveSlots': sum(1 for row in slot_rows if row['effectiveRuntimeState'] == 'live_now'),
        'reservedPending': sum(1 for row in slot_rows if row['effectiveRuntimeState'] == 'reserved_pending'),
        'activeCreatives': sum(1 for row in slot_rows if row['liveCreativeTitle'] != '-'),
        'lowCtrSlots': len(review_queues['lowCtr']),
        'noDataSlots': len(review_queues['noData']),
    }

    slot_rows.sort(key=lambda row: (0 if row['reviewFlag'] != 'ok' else 1, row['slotKey']))
    creative_rows.sort(key=lambda row: (-row['impressions'], row['title']))
    return {'summary': summary, 'slotRows': slot_rows, 'creativeRows': creative_rows, 'reviewQueues': review_queues}


def get_ad_slot_workspace(
    db: OrmSession,
    slot_key: str,
    *,
    period_from: Optional[str] = None,
    period_to: Optional[str] = None,
) -> Dict[str, Any]:
    slots = list_ad_slots(db)
    slot = next((row for row in slots if row.get('slotKey') == slot_key), None)
    if slot is None:
        raise ValueError('slot_not_found')

    history = list_ad_campaigns(db, slot_key=slot_key, limit=100)
    history_by_id = {row['id']: row for row in history}
    config = slot.get('config') or {}
    summary = get_ad_performance_summary(db, slot_key=slot_key, period_from=period_from, period_to=period_to)
    performance = summary['slotRows'][0] if summary['slotRows'] else None

    return {
        'slot': slot,
        'liveCampaign': history_by_id.get(config.get('liveCampaignId')),
        'reservedCampaign': history_by_id.get(config.get('reservedCampaignId')),
        'history': history,
        'performance': performance,
    }


def record_ad_impression(
    db: OrmSession,
    *,
    slot_key: str,
    campaign_id: str,
    screen_name: Optional[str] = None,
    user_id: Optional[str] = None,
    session_id: Optional[str] = None,
    creative_id: Optional[str] = None,
) -> Dict[str, Any]:
    ensure_default_ad_slots(db)
    slot = db.scalar(select(AdSlot).where(AdSlot.slot_key == slot_key))
    if slot is None:
        raise ValueError('slot_not_found')

    campaign = db.get(AdCampaign, campaign_id)
    if campaign is None or campaign.slot_id != slot.id:
        raise ValueError('campaign_not_found')

    impression = AdImpression(
        slot_id=slot.id,
        user_id=user_id,
        session_id=session_id,
        screen_name=screen_name,
        campaign_id=campaign.id,
        creative_id=creative_id or campaign.id,
    )
    db.add(impression)
    db.commit()
    db.refresh(impression)
    return {
        'impressionId': impression.id,
        'slotKey': slot.slot_key,
        'campaignId': campaign.id,
        'creativeId': impression.creative_id,
        'createdAt': impression.created_at.isoformat() if impression.created_at else None,
    }


def record_ad_click(db: OrmSession, *, impression_id: str) -> Dict[str, Any]:
    impression = db.get(AdImpression, impression_id)
    if impression is None:
        raise ValueError('impression_not_found')

    click = AdClick(impression_id=impression.id)
    db.add(click)
    db.commit()
    db.refresh(click)
    return {
        'clickId': click.id,
        'impressionId': impression.id,
        'campaignId': impression.campaign_id,
        'createdAt': click.created_at.isoformat() if click.created_at else None,
    }


def export_ad_campaign_csv(db: OrmSession, campaign_id: str) -> tuple[str, str]:
    campaign = db.get(AdCampaign, campaign_id)
    if campaign is None:
        raise ValueError('campaign_not_found')

    slot = db.get(AdSlot, campaign.slot_id)
    if slot is None:
        raise ValueError('slot_not_found')

    impressions = int(db.scalar(select(func.count(AdImpression.id)).where(AdImpression.campaign_id == campaign.id)) or 0)
    clicks = int(
        db.scalar(
            select(func.count(AdClick.id))
            .join(AdImpression, AdImpression.id == AdClick.impression_id)
            .where(AdImpression.campaign_id == campaign.id)
        )
        or 0
    )
    ctr = 0.0 if impressions == 0 else clicks / impressions

    rows = [
        ('campaign_id', campaign.id),
        ('slot_key', slot.slot_key),
        ('variant', campaign.variant),
        ('status', campaign.status),
        ('title', campaign.title),
        ('message', campaign.message),
        ('cta_label', campaign.cta_label or ''),
        ('target_url', campaign.target_url or ''),
        ('image_url', campaign.image_url or ''),
        ('audience_type', _campaign_audience_type(campaign)),
        ('target_region_level', _campaign_region_level(campaign)),
        ('target_city', _clean_region_string(campaign.target_city) or ''),
        ('target_district', _clean_region_string(campaign.target_district) or ''),
        ('target_neighborhood', _clean_region_string(campaign.target_neighborhood) or ''),
        ('target_region_keys', ', '.join(_campaign_region_keys(campaign))),
        ('target_label', _campaign_target_label(_campaign_audience_type(campaign), _campaign_region_level(campaign), _campaign_region_keys(campaign), _clean_region_string(campaign.target_city), _clean_region_string(campaign.target_district), _clean_region_string(campaign.target_neighborhood))),
        ('start_at', campaign.start_at.isoformat() if campaign.start_at else ''),
        ('end_at', campaign.end_at.isoformat() if campaign.end_at else ''),
        ('created_at', campaign.created_at.isoformat() if campaign.created_at else ''),
        ('updated_at', campaign.updated_at.isoformat() if campaign.updated_at else ''),
        ('impressions', impressions),
        ('clicks', clicks),
        ('ctr', round(ctr, 4)),
    ]

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['field', 'value'])
    writer.writerows(rows)
    filename = _safe_xlsx_filename(f'{slot.slot_key}-{campaign.variant}-{campaign.title or campaign.id}') + '.csv'
    return buffer.getvalue(), filename


def export_ad_campaign_xlsx(db: OrmSession, campaign_id: str) -> tuple[bytes, str]:
    campaign = db.get(AdCampaign, campaign_id)
    if campaign is None:
        raise ValueError('campaign_not_found')

    slot = db.get(AdSlot, campaign.slot_id)
    if slot is None:
        raise ValueError('slot_not_found')

    impressions = int(db.scalar(select(func.count(AdImpression.id)).where(AdImpression.campaign_id == campaign.id)) or 0)
    clicks = int(
        db.scalar(
            select(func.count(AdClick.id))
            .join(AdImpression, AdImpression.id == AdClick.impression_id)
            .where(AdImpression.campaign_id == campaign.id)
        )
        or 0
    )
    ctr = 0.0 if impressions == 0 else clicks / impressions

    workbook = Workbook()
    overview = workbook.active
    overview.title = 'Campaign'

    overview['A1'] = 'Cartly Ad Campaign Export'
    overview['A1'].font = Font(bold=True, size=16)

    rows = [
        ('campaign_id', campaign.id),
        ('slot_key', slot.slot_key),
        ('variant', campaign.variant),
        ('status', campaign.status),
        ('title', campaign.title),
        ('message', campaign.message),
        ('cta_label', campaign.cta_label or ''),
        ('target_url', campaign.target_url or ''),
        ('image_url', campaign.image_url or ''),
        ('audience_type', _campaign_audience_type(campaign)),
        ('target_region_level', _campaign_region_level(campaign)),
        ('target_city', _clean_region_string(campaign.target_city) or ''),
        ('target_district', _clean_region_string(campaign.target_district) or ''),
        ('target_neighborhood', _clean_region_string(campaign.target_neighborhood) or ''),
        ('target_region_keys', ', '.join(_campaign_region_keys(campaign))),
        ('target_label', _campaign_target_label(_campaign_audience_type(campaign), _campaign_region_level(campaign), _campaign_region_keys(campaign), _clean_region_string(campaign.target_city), _clean_region_string(campaign.target_district), _clean_region_string(campaign.target_neighborhood))),
        ('start_at', campaign.start_at.isoformat() if campaign.start_at else ''),
        ('end_at', campaign.end_at.isoformat() if campaign.end_at else ''),
        ('created_at', campaign.created_at.isoformat() if campaign.created_at else ''),
        ('updated_at', campaign.updated_at.isoformat() if campaign.updated_at else ''),
        ('impressions', impressions),
        ('clicks', clicks),
        ('ctr', round(ctr, 4)),
    ]
    for index, (label, value) in enumerate(rows, start=3):
        overview.cell(row=index, column=1, value=label).font = Font(bold=True)
        overview.cell(row=index, column=2, value=value)

    image_formula_url = _campaign_image_formula_url(campaign.image_url)
    if image_formula_url:
        overview['D3'] = 'image_preview'
        overview['D3'].font = Font(bold=True)
        overview['D4'] = f'=IMAGE("{image_formula_url}")'
        overview['D18'] = 'image_link'
        overview['D18'].font = Font(bold=True)
        overview['D19'] = image_formula_url
        overview.column_dimensions['D'].width = 28
        overview.column_dimensions['E'].width = 28
        overview.row_dimensions[4].height = 180
    else:
        overview['D3'] = 'image_preview'
        overview['D3'].font = Font(bold=True)
        overview['D4'] = '이미지 없음'

    for column in ['A', 'B']:
        overview.column_dimensions[column].width = 24 if column == 'A' else 42

    detail = workbook.create_sheet('Metrics')
    detail.append(['metric', 'value'])
    detail['A1'].font = Font(bold=True)
    detail['B1'].font = Font(bold=True)
    detail.append(['slot_key', slot.slot_key])
    detail.append(['campaign_id', campaign.id])
    detail.append(['variant', campaign.variant])
    detail.append(['status', campaign.status])
    detail.append(['impressions', impressions])
    detail.append(['clicks', clicks])
    detail.append(['ctr', round(ctr, 4)])
    detail.column_dimensions['A'].width = 22
    detail.column_dimensions['B'].width = 32

    output = io.BytesIO()
    workbook.save(output)
    filename = _safe_xlsx_filename(f'{slot.slot_key}-{campaign.variant}-{campaign.title or campaign.id}') + '.xlsx'
    return output.getvalue(), filename


def export_ad_campaigns_csv(
    db: OrmSession,
    *,
    slot_key: Optional[str] = None,
    query: Optional[str] = None,
    variant: Optional[str] = None,
    status: Optional[str] = None,
    period_from: Optional[str] = None,
    period_to: Optional[str] = None,
    limit: int = 500,
) -> tuple[str, str]:
    campaigns = list_ad_campaigns(
        db,
        slot_key=slot_key,
        limit=limit,
        query=query,
        variant=variant,
        status=status,
        period_from=period_from,
        period_to=period_to,
    )

    slot_query = select(AdSlot)
    if slot_key:
        slot_query = slot_query.where(AdSlot.slot_key == slot_key)
    slot_rows = db.scalars(slot_query).all()
    slot_map = {slot_row.slot_key: slot_row for slot_row in slot_rows}
    default_slot_map = {slot['slotKey']: slot for slot in DEFAULT_AD_SLOTS}

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([
        'slot_key',
        'slot_name',
        'slot_description',
        'placement_type',
        'screen',
        'position',
        'placement_note',
        'campaign_id',
        'variant',
        'status',
        'title',
        'message',
        'cta_label',
        'target_url',
        'image_url',
        'audience_type',
        'target_region_level',
        'target_city',
        'target_district',
        'target_neighborhood',
        'target_region_keys',
        'target_label',
        'start_at',
        'end_at',
        'impressions',
        'clicks',
        'ctr',
        'created_at',
        'updated_at',
    ])

    for campaign in campaigns:
        raw_slot_key = campaign.get('slotKey') or 'unknown'
        fallback_slot = default_slot_map.get(raw_slot_key)
        slot_config: Dict[str, Any] = fallback_slot['config'].copy() if fallback_slot else {}
        slot_row = slot_map.get(raw_slot_key)
        if slot_row and slot_row.config_json:
            try:
                slot_config.update(json.loads(slot_row.config_json) or {})
            except Exception:
                pass
        slot_config = _normalize_slot_config(slot_config, include_reserved=True)
        writer.writerow([
            raw_slot_key,
            slot_config.get('slotLabel') or raw_slot_key,
            slot_config.get('slotDescription') or '-',
            slot_row.placement_type if slot_row else fallback_slot['placementType'] if fallback_slot else '-',
            slot_config.get('screen') or '-',
            slot_config.get('position') or '-',
            slot_config.get('placementNote') or '-',
            campaign.get('id'),
            campaign.get('variant'),
            campaign.get('status'),
            campaign.get('title'),
            campaign.get('message'),
            campaign.get('ctaLabel'),
            campaign.get('targetUrl'),
            campaign.get('imageUrl'),
            campaign.get('audienceType'),
            campaign.get('targetRegionLevel'),
            campaign.get('targetCity'),
            campaign.get('targetDistrict'),
            campaign.get('targetNeighborhood'),
            ', '.join(campaign.get('targetRegionKeys') or []),
            campaign.get('targetLabel'),
            campaign.get('startAt'),
            campaign.get('endAt'),
            campaign.get('impressions'),
            campaign.get('clicks'),
            campaign.get('ctr'),
            campaign.get('createdAt'),
            campaign.get('updatedAt'),
        ])

    suffix = f'{period_from or "all"}_{period_to or "all"}'
    filename = _safe_xlsx_filename(f'cartly-ad-campaigns-{suffix}') + '.csv'
    return buffer.getvalue(), filename


def export_ad_campaigns_xlsx(
    db: OrmSession,
    *,
    slot_key: Optional[str] = None,
    query: Optional[str] = None,
    variant: Optional[str] = None,
    status: Optional[str] = None,
    period_from: Optional[str] = None,
    period_to: Optional[str] = None,
    limit: int = 500,
) -> tuple[bytes, str]:
    campaigns = list_ad_campaigns(
        db,
        slot_key=slot_key,
        limit=limit,
        query=query,
        variant=variant,
        status=status,
        period_from=period_from,
        period_to=period_to,
    )

    workbook = Workbook()
    overview = workbook.active
    overview.title = 'Overview'
    overview['A1'] = 'Cartly Past Ad Campaign Export'
    overview['A1'].font = Font(bold=True, size=16)

    overview_rows = [
        ('query', query or '-'),
        ('variant', variant or 'all'),
        ('status', status or 'all'),
        ('period_from', period_from or '-'),
        ('period_to', period_to or '-'),
        ('total_campaigns', len(campaigns)),
    ]
    for index, (label, value) in enumerate(overview_rows, start=3):
        overview.cell(row=index, column=1, value=label).font = Font(bold=True)
        overview.cell(row=index, column=2, value=value)
    overview.column_dimensions['A'].width = 24
    overview.column_dimensions['B'].width = 32

    headers = [
        'campaign_id',
        'variant',
        'status',
        'title',
        'message',
        'cta_label',
        'target_url',
        'image_url',
        'image_preview',
        'audience_type',
        'target_region_level',
        'target_city',
        'target_district',
        'target_neighborhood',
        'target_region_keys',
        'target_label',
        'start_at',
        'end_at',
        'impressions',
        'clicks',
        'ctr',
        'created_at',
        'updated_at',
    ]
    slot_query = select(AdSlot)
    if slot_key:
        slot_query = slot_query.where(AdSlot.slot_key == slot_key)
    slot_rows = db.scalars(slot_query).all()
    slot_map = {slot_row.slot_key: slot_row for slot_row in slot_rows}
    default_slot_map = {slot['slotKey']: slot for slot in DEFAULT_AD_SLOTS}

    campaigns_by_slot: Dict[str, List[Dict[str, Any]]] = {}
    for campaign in campaigns:
        campaigns_by_slot.setdefault(campaign.get('slotKey') or 'unknown', []).append(campaign)

    for raw_slot_key, slot_campaigns in campaigns_by_slot.items():
        title = _safe_xlsx_filename(raw_slot_key).replace('-', '_')[:31] or 'slot'
        sheet = workbook.create_sheet(title)

        fallback_slot = default_slot_map.get(raw_slot_key)
        slot_config: Dict[str, Any] = fallback_slot['config'].copy() if fallback_slot else {}
        slot_row = slot_map.get(raw_slot_key)
        if slot_row and slot_row.config_json:
            try:
                slot_config.update(json.loads(slot_row.config_json) or {})
            except Exception:
                pass
        slot_config = _normalize_slot_config(slot_config, include_reserved=True)

        meta_rows = [
            ('slot_key', raw_slot_key),
            ('slot_name', slot_config.get('slotLabel') or raw_slot_key),
            ('slot_description', slot_config.get('slotDescription') or '-'),
            ('placement_type', slot_row.placement_type if slot_row else fallback_slot['placementType'] if fallback_slot else '-'),
            ('screen', slot_config.get('screen') or '-'),
            ('position', slot_config.get('position') or '-'),
            ('placement_note', slot_config.get('placementNote') or '-'),
            ('total_campaigns', len(slot_campaigns)),
        ]
        for index, (label, value) in enumerate(meta_rows, start=1):
            sheet.cell(row=index, column=1, value=label).font = Font(bold=True)
            sheet.cell(row=index, column=2, value=value)

        table_start_row = len(meta_rows) + 3
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=table_start_row, column=col_index, value=header).font = Font(bold=True)

        for row_offset, campaign in enumerate(slot_campaigns, start=1):
            image_url = _campaign_image_formula_url(campaign.get('imageUrl'))
            values = [
                campaign.get('id'),
                campaign.get('variant'),
                campaign.get('status'),
                campaign.get('title'),
                campaign.get('message'),
                campaign.get('ctaLabel'),
                campaign.get('targetUrl'),
                campaign.get('imageUrl'),
                f'=IMAGE("{image_url}")' if image_url else '',
                campaign.get('audienceType'),
                campaign.get('targetRegionLevel'),
                campaign.get('targetCity'),
                campaign.get('targetDistrict'),
                campaign.get('targetNeighborhood'),
                ', '.join(campaign.get('targetRegionKeys') or []),
                campaign.get('targetLabel'),
                campaign.get('startAt'),
                campaign.get('endAt'),
                campaign.get('impressions'),
                campaign.get('clicks'),
                campaign.get('ctr'),
                campaign.get('createdAt'),
                campaign.get('updatedAt'),
            ]
            for col_index, value in enumerate(values, start=1):
                sheet.cell(row=table_start_row + row_offset, column=col_index, value=value)

        sheet.freeze_panes = f'A{table_start_row + 1}'
        for column in sheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 36)

    output = io.BytesIO()
    workbook.save(output)
    suffix = f'{period_from or "all"}_{period_to or "all"}'
    filename = _safe_xlsx_filename(f'cartly-ad-campaigns-{suffix}') + '.xlsx'
    return output.getvalue(), filename


def app_ad_slots_config(
    db: OrmSession,
    ads_enabled: bool,
    *,
    current_user: Optional[User] = None,
    city: Optional[str] = None,
    district: Optional[str] = None,
    neighborhood: Optional[str] = None,
) -> List[Dict[str, Any]]:
    ensure_default_ad_slots(db)
    rows = db.scalars(select(AdSlot).order_by(AdSlot.created_at.asc())).all()
    defaults = _defaults_by_key()
    dirty = False
    for row in rows:
        if row.slot_key in defaults:
            before = row.config_json
            _sync_slot_runtime_config(db, row, defaults[row.slot_key])
            if row.config_json != before:
                dirty = True
    if dirty:
        db.commit()
        for row in rows:
            db.refresh(row)

    slot_ids = [row.id for row in rows if row.slot_key in defaults]
    campaign_rows = db.scalars(
        select(AdCampaign).where(AdCampaign.slot_id.in_(slot_ids)).order_by(AdCampaign.created_at.asc())
    ).all() if slot_ids else []
    campaigns_by_slot: Dict[str, List[AdCampaign]] = {}
    for campaign in campaign_rows:
        campaigns_by_slot.setdefault(campaign.slot_id, []).append(campaign)

    runtime_context = build_ad_runtime_context(
        current_user,
        city=city,
        district=district,
        neighborhood=neighborhood,
    )

    result = []
    for row in rows:
        fallback = defaults.get(row.slot_key)
        if fallback is None:
            continue
        projected = _project_slot_config_from_campaigns(
            _base_slot_config(row, fallback),
            campaigns_by_slot.get(row.id, []),
            target_context=runtime_context,
        )
        live_creatives = list(projected.get('liveCreatives') or [])
        enabled = ads_enabled and row.status == 'active' and len(live_creatives) > 0
        result.append(
            {
                'slotKey': row.slot_key,
                'placementType': row.placement_type,
                'enabled': enabled,
                'config': {
                    **_normalize_slot_config(projected),
                    'campaignId': projected.get('liveCampaignId'),
                    'creatives': live_creatives,
                    'rotationMode': projected.get('liveRotationMode') or ('single' if len(live_creatives) <= 1 else 'ordered'),
                    'landingType': projected.get('landingType'),
                    'landingKey': projected.get('landingKey'),
                    'landingParams': projected.get('landingParams'),
                },
            }
        )
    return result
